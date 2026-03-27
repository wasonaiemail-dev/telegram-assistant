"""
alfred/features/workout.py
===========================
Full workout tracking, GPT-generated programs, PR logging, Excel export.

STORAGE
───────
  workout_program.json  (WORKOUT_FILE) — program, templates, PRs, body stats
  workout_log.xlsx      (WORKOUT_XLSX) — every logged session

COMMANDS
────────
  /workout           — view today's program day + recent streak
  /workout export    — send the workout log Excel file

INTENTS HANDLED
───────────────
  WORKOUT_LOG, WORKOUT_VIEW, WORKOUT_ASK, WORKOUT_PLAN, WORKOUT_REBUILD,
  WORKOUT_TEMPLATE, WORKOUT_EXPORT, WORKOUT_BODY
"""

import os
import re
import json
import logging
import datetime

import pytz

from core.config import (
    BOT_NAME, TIMEZONE, OPENAI_API_KEY, GPT_CHAT_MODEL,
    WORKOUT_XLSX, WORKOUT_FILE,
)
from core.intent import (
    WORKOUT_LOG, WORKOUT_VIEW, WORKOUT_ASK, WORKOUT_PLAN,
    WORKOUT_REBUILD, WORKOUT_TEMPLATE, WORKOUT_EXPORT, WORKOUT_BODY,
)
from core.data import load_workout, save_workout

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL LOG HELPERS
# ─────────────────────────────────────────────────────────────────────────────

_LOG_COLS = [
    "date", "day_label", "duration_min", "energy_1_5",
    "exercise", "sets", "reps", "weight_lb",
    "cardio_distance_km", "cardio_pace_min_km", "notes",
]


def _ensure_log_xlsx():
    if os.path.exists(WORKOUT_XLSX):
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workout Log"
        fill = PatternFill("solid", fgColor="7C3AED")
        for col_idx, col_name in enumerate(_LOG_COLS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")
        wb.save(WORKOUT_XLSX)
        logger.info("Created workout_log.xlsx")
    except ImportError:
        logger.warning("openpyxl not installed — workout Excel unavailable")


def _append_session_to_xlsx(session: dict, exercises: list[dict], date_iso: str, day_label: str = ""):
    """Write one workout session (potentially multiple exercise rows) to the log."""
    _ensure_log_xlsx()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(WORKOUT_XLSX)
        ws = wb["Workout Log"]
        dur    = session.get("duration_min", "")
        energy = session.get("energy", "")
        cardio = session.get("cardio", {})
        dist   = cardio.get("distance_km", "")
        pace   = cardio.get("pace_min_km", "")
        notes  = session.get("notes", "")

        if exercises:
            for ex in exercises:
                ws.append([
                    date_iso, day_label, dur, energy,
                    ex.get("exercise", ""), ex.get("sets", ""),
                    ex.get("reps", ""), ex.get("weight_lb", ""),
                    dist, pace, notes,
                ])
        else:
            # Cardio or no-exercise session
            ws.append([date_iso, day_label, dur, energy, "", "", "", "", dist, pace, notes])

        wb.save(WORKOUT_XLSX)
        return True
    except Exception as e:
        logger.error(f"Error saving workout to xlsx: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# GPT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

async def _gpt_generate_program(settings: dict) -> dict:
    """Generate a weekly workout program from the buyer's setup settings."""
    from openai import AsyncOpenAI
    client = AsyncOpenAI(api_key=OPENAI_API_KEY)
    prompt = f"""Generate a weekly workout program with these parameters:
Goal: {settings.get('goal', 'general fitness')}
Days per week: {settings.get('days_per_week', 4)}
Equipment: {settings.get('equipment', 'gym')}
Preferences / restrictions: {settings.get('preferences', 'none')}
Progressive overload: {'yes' if settings.get('progressive_overload', True) else 'no'}

Return ONLY valid JSON — a dict with day labels as keys (e.g. "Day 1 - Push", "Day 2 - Pull").
Each value is a list of exercise objects: {{exercise, sets (int), reps (str e.g. "8-10"), weight_note (str e.g. "moderate")}}
Include rest days as {{"rest": true}}.
No markdown, no extra text."""
    try:
        resp = await client.chat.completions.create(
            model=GPT_CHAT_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
        )
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r'^```json\s*|\s*```$', '', raw, flags=re.DOTALL)
        return json.loads(raw)
    except Exception as e:
        logger.error(f"GPT program generation failed: {e}")
        return {}


async def _gpt_suggest_workout(w: dict, muscle_group: str = "") -> str:
    """Suggest today's workout based on the program and recent history."""
    from openai import AsyncOpenAI
    client = AsyncOpenAI(api_key=OPENAI_API_KEY)
    program  = w.get("program", {})
    pr_log   = w.get("pr_log", {})
    settings = {
        "goal":      w.get("goal", "general fitness"),
        "equipment": w.get("equipment", "gym"),
        "progressive_overload": w.get("progressive_overload", True),
    }
    last_workout = w.get("last_workout_date")
    streak       = w.get("streak", 0)

    context_text = f"""
Workout program: {json.dumps(program)}
Personal records: {json.dumps(pr_log)}
Last workout date: {last_workout or 'N/A'}
Current streak: {streak} days
Progressive overload enabled: {settings['progressive_overload']}
Goal: {settings['goal']}
Equipment: {settings['equipment']}
{f"Requested muscle group: {muscle_group}" if muscle_group else ""}
""".strip()

    prompt = f"""Based on this user's workout data, suggest today's workout session.

{context_text}

Include:
1. Session name / focus
2. Each exercise with specific sets, reps, and weight (accounting for progressive overload from PRs)
3. A brief motivational note

Keep it concise and practical. Use Markdown bold for exercise names."""
    try:
        resp = await client.chat.completions.create(
            model=GPT_CHAT_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.6,
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        logger.error(f"GPT workout suggestion failed: {e}")
        return "Couldn't generate a suggestion right now. Try again!"


def _parse_workout_log(description: str) -> tuple[list[dict], dict]:
    """
    Parse a natural-language post-workout log into exercises and cardio.
    Returns (exercises, cardio_dict).
    Example: "bench 3x8 at 185, squats 4x6 at 225, ran 5km in 28 min"
    """
    exercises = []
    cardio    = {}

    # Cardio: "ran/walked/cycled Xkm in Y min" or "X km at Y pace"
    cardio_pattern = re.compile(
        r'(?:ran|run|walked|walk|cycled|cycle|biked|bike|swam|swim|hiked|hike)\s+'
        r'(\d+(?:\.\d+)?)\s*km\s+(?:in\s+(\d+)\s*min)?',
        re.IGNORECASE,
    )
    m = cardio_pattern.search(description)
    if m:
        dist_km = float(m.group(1))
        mins    = int(m.group(2)) if m.group(2) else None
        cardio["distance_km"] = dist_km
        if mins:
            pace = round(mins / dist_km, 2)
            cardio["pace_min_km"] = pace
        description = description[:m.start()] + description[m.end():]

    # Strength: "exercise NxM at Wlb" or "exercise N sets M reps W"
    ex_pattern = re.compile(
        r'([A-Za-z ]+?)\s+'
        r'(\d+)\s*[x×]\s*(\d+)'
        r'(?:\s+(?:at|@)\s*(\d+(?:\.\d+)?)\s*(?:lb|lbs|kg)?)?',
        re.IGNORECASE,
    )
    for m in ex_pattern.finditer(description):
        name       = m.group(1).strip().title()
        sets       = int(m.group(2))
        reps       = int(m.group(3))
        weight_lb  = float(m.group(4)) if m.group(4) else None
        exercises.append({
            "exercise":  name,
            "sets":      sets,
            "reps":      reps,
            "weight_lb": weight_lb,
        })

    return exercises, cardio


def _check_prs(exercises: list[dict], w: dict) -> list[str]:
    """Check for new PRs and update pr_log. Returns list of PR announcement strings."""
    prs       = []
    pr_log    = w.setdefault("pr_log", {})
    for ex in exercises:
        name   = ex.get("exercise", "")
        weight = ex.get("weight_lb")
        reps   = ex.get("reps", 0)
        if not name or not weight:
            continue
        key    = name.lower()
        old_pr = pr_log.get(key, {})
        old_w  = old_pr.get("weight_lb", 0) or 0
        if weight > old_w:
            pr_log[key] = {
                "weight_lb": weight,
                "reps":      reps,
                "date":      datetime.date.today().isoformat(),
            }
            prs.append(f"🏆 New PR: *{name}* — {weight} lb × {reps} reps!")
    return prs


def _update_streak(w: dict, today_iso: str) -> int:
    """Update workout streak. Returns new streak count."""
    last = w.get("last_workout_date")
    if last:
        last_date  = datetime.date.fromisoformat(last)
        today_date = datetime.date.fromisoformat(today_iso)
        delta      = (today_date - last_date).days
        if delta == 1:
            w["streak"] = w.get("streak", 0) + 1
        elif delta > 1:
            w["streak"] = 1
        # delta == 0: same day, no change
    else:
        w["streak"] = 1
    w["last_workout_date"] = today_iso
    return w["streak"]


# ─────────────────────────────────────────────────────────────────────────────
# BRIEFING HELPER
# ─────────────────────────────────────────────────────────────────────────────

def get_briefing_line() -> str:
    """One-line workout summary for morning briefing."""
    try:
        w = load_workout()
        streak  = w.get("streak", 0)
        last    = w.get("last_workout_date")
        if not last:
            return ""
        last_date  = datetime.date.fromisoformat(last)
        today_date = datetime.date.today()
        days_ago   = (today_date - last_date).days
        if days_ago > 7:
            return ""
        parts = [f"💪 *Workout Streak:* {streak} day{'s' if streak != 1 else ''}"]
        if days_ago == 0:
            parts.append("(worked out today)")
        elif days_ago == 1:
            parts.append("(last workout: yesterday)")
        else:
            parts.append(f"(last workout: {days_ago} days ago)")
        return " ".join(parts)
    except Exception:
        return ""


# ─────────────────────────────────────────────────────────────────────────────
# /workout COMMAND
# ─────────────────────────────────────────────────────────────────────────────

async def cmd_workout(update, context) -> None:
    """Show workout program overview and streak."""
    w = load_workout()
    program = w.get("program", {})
    streak  = w.get("streak", 0)
    last    = w.get("last_workout_date", "N/A")

    lines = [f"💪 *Workout Overview*", f"Streak: {streak} day{'s' if streak != 1 else ''}  |  Last: {last}"]

    if program:
        lines.append("\n*Your Program:*")
        for day_label, exercises in list(program.items())[:7]:
            if isinstance(exercises, dict) and exercises.get("rest"):
                lines.append(f"  {day_label}: Rest")
            elif isinstance(exercises, list):
                ex_names = ", ".join(e.get("exercise", "") for e in exercises[:3])
                if len(exercises) > 3:
                    ex_names += f" +{len(exercises)-3} more"
                lines.append(f"  {day_label}: {ex_names}")
    else:
        lines.append("\nNo program set. Say \"build my workout program\" to generate one.")

    lines.append('\nSay "what should I do today?" for a personalised suggestion.')
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


# ─────────────────────────────────────────────────────────────────────────────
# INTENT HANDLER
# ─────────────────────────────────────────────────────────────────────────────

async def handle_workout_intent(intent: str, entities: dict, update, context) -> None:
    msg = update.message
    tz  = pytz.timezone(TIMEZONE)
    today_iso = datetime.datetime.now(tz).date().isoformat()

    # ── WORKOUT_LOG ───────────────────────────────────────────────────────────
    if intent == WORKOUT_LOG:
        description  = entities.get("description", "")
        duration_min = entities.get("duration_min")
        energy       = entities.get("energy")
        exercises    = entities.get("exercises") or []
        cardio_ent   = entities.get("cardio", {})

        # Parse description if no structured exercises given
        if not exercises and description:
            parsed_ex, parsed_cardio = _parse_workout_log(description)
            if parsed_ex:
                exercises = parsed_ex
            if parsed_cardio and not cardio_ent:
                cardio_ent = parsed_cardio

        if not exercises and not cardio_ent and not duration_min:
            await msg.reply_text(
                "I didn't catch what you did. Try:\n"
                "\"bench 3x8 at 185, squats 4x6 at 225, ran 3km in 20 min — 45 min session, energy 4\""
            )
            return

        w = load_workout()
        # Check PRs
        pr_msgs = _check_prs(exercises, w)
        # Update streak
        new_streak = _update_streak(w, today_iso)
        # Get current program day label
        day_label = ""
        program   = w.get("program", {})
        if program:
            day_keys  = [k for k in program if not (isinstance(program[k], dict) and program[k].get("rest"))]
            day_label = day_keys[(new_streak - 1) % max(len(day_keys), 1)] if day_keys else ""

        save_workout(w)

        session = {
            "duration_min": duration_min,
            "energy":       energy,
            "cardio":       cardio_ent,
            "notes":        description,
        }
        _append_session_to_xlsx(session, exercises, today_iso, day_label)

        lines = [f"✓ Workout logged! Streak: {new_streak} day{'s' if new_streak != 1 else ''}"]
        if exercises:
            lines.append(f"  {len(exercises)} exercise{'s' if len(exercises) != 1 else ''} recorded")
        if cardio_ent:
            d = cardio_ent.get('distance_km', '')
            p = cardio_ent.get('pace_min_km', '')
            lines.append(f"  Cardio: {d}km" + (f" @ {p} min/km" if p else ""))
        if pr_msgs:
            lines.extend(pr_msgs)

        await msg.reply_text("\n".join(lines), parse_mode="Markdown")
        return

    # ── WORKOUT_VIEW ──────────────────────────────────────────────────────────
    if intent == WORKOUT_VIEW:
        days = int(entities.get("days", 7))
        _ensure_log_xlsx()
        if not os.path.exists(WORKOUT_XLSX):
            await msg.reply_text("No workout history yet. Log a session to get started!")
            return
        try:
            import openpyxl
            wb   = openpyxl.load_workbook(WORKOUT_XLSX, read_only=True)
            ws   = wb["Workout Log"]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                await msg.reply_text("No workout sessions logged yet.")
                return
            cutoff = (datetime.date.today() - datetime.timedelta(days=days)).isoformat()
            recent = [r for r in rows[1:] if r[0] and str(r[0]) >= cutoff]
            if not recent:
                await msg.reply_text(f"No workouts in the last {days} days.")
                return
            lines = [f"💪 *Last {days} days ({len(recent)} session{'s' if len(recent)!=1 else ''})*"]
            seen_dates = set()
            for row in reversed(recent[-10:]):
                date = str(row[0])
                if date not in seen_dates:
                    seen_dates.add(date)
                    label = str(row[1] or "")
                    dur   = f"{row[2]} min" if row[2] else ""
                    lines.append(f"  {date}: {label} {dur}".strip())
            await msg.reply_text("\n".join(lines), parse_mode="Markdown")
        except Exception as e:
            logger.error(f"Error reading workout log: {e}")
            await msg.reply_text("Couldn't read workout history. Try again.")
        return

    # ── WORKOUT_ASK (on-demand suggestion) ────────────────────────────────────
    if intent == WORKOUT_ASK:
        muscle = entities.get("muscle_group", "")
        await msg.reply_text("⏳ Coming up with today's session...")
        w       = load_workout()
        suggest = await _gpt_suggest_workout(w, muscle_group=muscle)
        await msg.reply_text(suggest, parse_mode="Markdown")
        return

    # ── WORKOUT_PLAN (view) ───────────────────────────────────────────────────
    if intent == WORKOUT_PLAN:
        w = load_workout()
        program = w.get("program", {})
        if not program:
            await msg.reply_text("No program yet. Say \"build my workout program\" to generate one.")
            return
        lines = ["💪 *Your Workout Program*"]
        for day_label, exercises in program.items():
            if isinstance(exercises, dict) and exercises.get("rest"):
                lines.append(f"\n*{day_label}* — Rest day")
            elif isinstance(exercises, list):
                lines.append(f"\n*{day_label}*")
                for ex in exercises:
                    weight_note = f" ({ex.get('weight_note','')})" if ex.get('weight_note') else ""
                    lines.append(f"  • {ex.get('exercise','')} — {ex.get('sets','')}×{ex.get('reps','')}{weight_note}")
        await msg.reply_text("\n".join(lines), parse_mode="Markdown")
        return

    # ── WORKOUT_REBUILD ───────────────────────────────────────────────────────
    if intent == WORKOUT_REBUILD:
        from core.data import load_data, get_workout_settings
        data     = load_data()
        settings = get_workout_settings(data)
        await msg.reply_text("⏳ Rebuilding your workout program...")
        program = await _gpt_generate_program(settings)
        if not program:
            await msg.reply_text("Couldn't generate a program right now. Try again.")
            return
        w = load_workout()
        w["program"] = program
        w.update({
            "goal":               settings.get("goal", w.get("goal")),
            "days_per_week":      settings.get("days_per_week", w.get("days_per_week")),
            "equipment":          settings.get("equipment", w.get("equipment")),
            "preferences":        settings.get("preferences", w.get("preferences")),
            "progressive_overload": settings.get("progressive_overload", w.get("progressive_overload")),
        })
        save_workout(w)
        day_labels = list(program.keys())
        await msg.reply_text(
            f"✓ New {len(day_labels)}-day program created:\n" +
            "\n".join(f"  • {d}" for d in day_labels),
            parse_mode="Markdown",
        )
        return

    # ── WORKOUT_TEMPLATE ──────────────────────────────────────────────────────
    if intent == WORKOUT_TEMPLATE:
        action = entities.get("action", "list")
        name   = entities.get("name", "").strip()
        w      = load_workout()
        templates = w.setdefault("templates", {})

        if action == "list":
            if not templates:
                await msg.reply_text("No saved templates yet. Say \"save this as [name]\" after logging a workout.")
                return
            lines = ["💾 *Saved Templates:*"] + [f"  • {n}" for n in templates]
            await msg.reply_text("\n".join(lines), parse_mode="Markdown")

        elif action == "save":
            if not name:
                await msg.reply_text("What should I call this template?")
                return
            # Save last logged day's exercises from the xlsx as a template
            _ensure_log_xlsx()
            try:
                import openpyxl
                wb   = openpyxl.load_workbook(WORKOUT_XLSX, read_only=True)
                ws   = wb["Workout Log"]
                rows = list(ws.iter_rows(values_only=True))
                last_date = str(rows[-1][0]) if len(rows) > 1 else ""
                template_exercises = [
                    {"exercise": str(r[4]), "sets": r[5], "reps": r[6], "weight_lb": r[7]}
                    for r in rows[1:] if str(r[0]) == last_date and r[4]
                ]
                templates[name] = template_exercises
                save_workout(w)
                await msg.reply_text(f"✓ Template \"{name}\" saved with {len(template_exercises)} exercises.")
            except Exception as e:
                await msg.reply_text("Couldn't save template from last session. Try again.")

        elif action == "load":
            if not name or name not in templates:
                await msg.reply_text(
                    f"Template \"{name}\" not found." if name else "Which template? " +
                    ", ".join(f"\"{n}\"" for n in templates)
                )
                return
            exercises = templates[name]
            lines = [f"📋 *{name}:*"]
            for ex in exercises:
                pr_note = ""
                pr_log  = w.get("pr_log", {})
                pr      = pr_log.get(str(ex.get("exercise","")).lower())
                if pr and w.get("progressive_overload"):
                    suggested = (pr.get("weight_lb") or 0) + 5
                    pr_note   = f" → suggest {suggested} lb today"
                lines.append(f"  • {ex.get('exercise','')} — {ex.get('sets','')}×{ex.get('reps','')}{pr_note}")
            await msg.reply_text("\n".join(lines), parse_mode="Markdown")
        return

    # ── WORKOUT_EXPORT ────────────────────────────────────────────────────────
    if intent == WORKOUT_EXPORT:
        _ensure_log_xlsx()
        if not os.path.exists(WORKOUT_XLSX):
            await msg.reply_text("No workout history yet.")
            return
        await context.bot.send_document(
            chat_id=msg.chat_id,
            document=open(WORKOUT_XLSX, "rb"),
            filename="workout_log.xlsx",
            caption="Your complete workout log.",
        )
        return

    # ── WORKOUT_BODY ──────────────────────────────────────────────────────────
    if intent == WORKOUT_BODY:
        action       = entities.get("action", "log")
        weight_lb    = entities.get("weight_lb")
        measurements = entities.get("measurements", {})
        w = load_workout()

        if action == "log":
            entry = {
                "date":         today_iso,
                "weight_lb":    weight_lb,
                "measurements": measurements,
            }
            w.setdefault("body_stats", []).append(entry)
            save_workout(w)
            parts = []
            if weight_lb:
                parts.append(f"weight: {weight_lb} lb")
            if measurements:
                parts.extend(f"{k}: {v}" for k, v in measurements.items())
            await msg.reply_text(f"✓ Body stats logged: {', '.join(parts) if parts else 'entry saved'}.")

        elif action == "view":
            stats = w.get("body_stats", [])
            if not stats:
                await msg.reply_text("No body stats logged yet.")
                return
            recent = stats[-5:]
            lines  = ["📏 *Recent Body Stats:*"]
            for s in reversed(recent):
                line = f"  {s['date']}"
                if s.get("weight_lb"):
                    line += f"  |  {s['weight_lb']} lb"
                if s.get("measurements"):
                    for k, v in s["measurements"].items():
                        line += f"  |  {k}: {v}"
                lines.append(line)
            await msg.reply_text("\n".join(lines), parse_mode="Markdown")
        return
