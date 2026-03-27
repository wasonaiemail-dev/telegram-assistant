"""
alfred/features/meals.py
========================
Full meal planning, recipe library, nutrition tracking, and meal adherence.

STORAGE
───────
  meals.xlsx  (MEALS_XLSX in config)
    Sheet "Recipes"   — recipe library (name, category, prep_min, cook_min,
                         servings, calories, protein_g, carbs_g, fat_g,
                         ingredients, instructions, rating, tags)
    Sheet "Meal Plan" — weekly planner (date, breakfast, lunch, dinner, snacks)
    Sheet "Adherence" — what was actually eaten vs planned

  userdata.json["settings"]["meal_adherence_time"]
    Evening adherence check reminder time (HH:MM)

COMMANDS
────────
  /meals             — show today's planned meals
  /meals plan [date] — view meal plan for a date / week
  /meals recipe [name] — look up a recipe
  /meals export      — send the meals Excel file

INTENTS HANDLED
───────────────
  MEAL_PLAN, MEAL_VIEW, MEAL_ADD, MEAL_RECIPE, MEAL_GENERATE,
  MEAL_IMPORT, MEAL_NUTRITION, MEAL_ADHERENCE, MEAL_EXPORT, MEAL_LEFTOVERS
"""

import os
import re
import json
import asyncio
import logging
import datetime

import pytz

from core.config import BOT_NAME, TIMEZONE, MEALS_XLSX, OPENAI_API_KEY, GPT_CHAT_MODEL
from core.intent import (
    MEAL_PLAN, MEAL_VIEW, MEAL_ADD, MEAL_RECIPE, MEAL_GENERATE,
    MEAL_IMPORT, MEAL_NUTRITION, MEAL_ADHERENCE, MEAL_EXPORT, MEAL_LEFTOVERS,
)

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

_RECIPE_COLS = [
    "name", "category", "prep_min", "cook_min", "servings",
    "calories", "protein_g", "carbs_g", "fat_g",
    "ingredients", "instructions", "rating", "tags",
]
_PLAN_COLS    = ["date", "breakfast", "lunch", "dinner", "snacks", "notes"]
_ADHERE_COLS  = ["date", "meal_slot", "planned", "actual", "notes"]


def _ensure_workbook():
    """Create meals.xlsx with required sheets if it doesn't exist."""
    if os.path.exists(MEALS_XLSX):
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()

        def _make_sheet(wb, title, cols, accent):
            if title == "Sheet":
                ws = wb.active
                ws.title = title
            else:
                ws = wb.create_sheet(title)
            header_fill = PatternFill("solid", fgColor=accent)
            for col_idx, col_name in enumerate(cols, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center")
            return ws

        _make_sheet(wb, "Recipes",    _RECIPE_COLS, "2563EB")
        _make_sheet(wb, "Meal Plan",  _PLAN_COLS,   "059669")
        _make_sheet(wb, "Adherence",  _ADHERE_COLS, "D97706")

        # Remove default "Sheet" tab if it was auto-created
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(MEALS_XLSX)
        logger.info("Created meals.xlsx")
    except ImportError:
        logger.warning("openpyxl not installed — meals Excel unavailable")


def _load_recipes() -> list[dict]:
    """Read all recipes from the Recipes sheet."""
    _ensure_workbook()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(MEALS_XLSX, read_only=True)
        ws = wb["Recipes"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(c).lower().replace(" ", "_") for c in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(row)]
    except Exception as e:
        logger.error(f"Error reading recipes: {e}")
        return []


def _save_recipe(recipe: dict) -> bool:
    """Append a recipe row to the Recipes sheet."""
    _ensure_workbook()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(MEALS_XLSX)
        ws = wb["Recipes"]
        row = [recipe.get(c, "") for c in _RECIPE_COLS]
        ws.append(row)
        wb.save(MEALS_XLSX)
        return True
    except Exception as e:
        logger.error(f"Error saving recipe: {e}")
        return False


def _get_plan_for_date(date_iso: str) -> dict | None:
    """Return the meal plan row for a specific date, or None."""
    _ensure_workbook()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(MEALS_XLSX, read_only=True)
        ws = wb["Meal Plan"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return None
        headers = [str(c).lower().replace(" ", "_") for c in rows[0]]
        for row in rows[1:]:
            d = dict(zip(headers, row))
            if str(d.get("date", "")) == date_iso:
                return d
        return None
    except Exception as e:
        logger.error(f"Error reading meal plan: {e}")
        return None


def _set_plan_for_date(date_iso: str, meals: dict) -> bool:
    """Set or update the meal plan for a date."""
    _ensure_workbook()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(MEALS_XLSX)
        ws = wb["Meal Plan"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(c).lower().replace(" ", "_") for c in rows[0]] if rows else _PLAN_COLS

        # Find existing row
        existing_row = None
        for idx, row in enumerate(rows[1:], 2):
            if str(row[0]) == date_iso:
                existing_row = idx
                break

        new_row = [date_iso] + [meals.get(c, "") for c in _PLAN_COLS[1:]]
        if existing_row:
            for col_idx, val in enumerate(new_row, 1):
                ws.cell(row=existing_row, column=col_idx, value=val)
        else:
            ws.append(new_row)
        wb.save(MEALS_XLSX)
        return True
    except Exception as e:
        logger.error(f"Error saving meal plan: {e}")
        return False


def _log_adherence(date_iso: str, slot: str, planned: str, actual: str, notes: str = "") -> bool:
    """Record actual vs planned meal."""
    _ensure_workbook()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(MEALS_XLSX)
        ws = wb["Adherence"]
        ws.append([date_iso, slot, planned, actual, notes])
        wb.save(MEALS_XLSX)
        return True
    except Exception as e:
        logger.error(f"Error logging adherence: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _today_iso() -> str:
    tz  = pytz.timezone(TIMEZONE)
    return datetime.datetime.now(tz).date().isoformat()


def _find_recipe(name: str) -> dict | None:
    """Find a recipe by name (case-insensitive partial match)."""
    name_lower = name.lower()
    recipes    = _load_recipes()
    for r in recipes:
        if name_lower in str(r.get("name", "")).lower():
            return r
    return None


def _format_recipe(r: dict) -> str:
    lines = [f"🍽 *{r.get('name', 'Recipe')}*"]
    if r.get("category"):
        lines.append(f"Category: {r['category']}")
    timing = []
    if r.get("prep_min"):
        timing.append(f"Prep: {r['prep_min']} min")
    if r.get("cook_min"):
        timing.append(f"Cook: {r['cook_min']} min")
    if timing:
        lines.append(" · ".join(timing))
    if r.get("servings"):
        lines.append(f"Servings: {r['servings']}")
    if any(r.get(k) for k in ("calories", "protein_g", "carbs_g", "fat_g")):
        macros = []
        if r.get("calories"):
            macros.append(f"🔥 {r['calories']} cal")
        if r.get("protein_g"):
            macros.append(f"💪 {r['protein_g']}g protein")
        if r.get("carbs_g"):
            macros.append(f"🌾 {r['carbs_g']}g carbs")
        if r.get("fat_g"):
            macros.append(f"🥑 {r['fat_g']}g fat")
        lines.append(" · ".join(macros))
    if r.get("ingredients"):
        lines.append(f"\n*Ingredients:*\n{r['ingredients']}")
    if r.get("instructions"):
        lines.append(f"\n*Instructions:*\n{r['instructions']}")
    return "\n".join(lines)


def _format_day_plan(plan: dict) -> str:
    lines = [f"📅 *Meals for {plan.get('date', 'today')}*"]
    for slot in ("breakfast", "lunch", "dinner", "snacks"):
        val = plan.get(slot)
        if val:
            lines.append(f"  {slot.title()}: {val}")
    if plan.get("notes"):
        lines.append(f"  _Note: {plan['notes']}_")
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# GPT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

async def _gpt_generate_recipe(description: str) -> dict | None:
    """Use GPT to generate a recipe from a description."""
    from openai import AsyncOpenAI
    client = AsyncOpenAI(api_key=OPENAI_API_KEY)
    prompt = f"""Generate a recipe for: {description}

Return ONLY valid JSON with these keys:
name, category, prep_min (int), cook_min (int), servings (int),
calories (int per serving), protein_g (int), carbs_g (int), fat_g (int),
ingredients (multiline string), instructions (numbered multiline string), tags (comma list)

No markdown, no extra text — just the JSON object."""
    try:
        resp = await client.chat.completions.create(
            model=GPT_CHAT_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
        )
        raw  = resp.choices[0].message.content.strip()
        raw  = re.sub(r'^```json\s*|\s*```$', '', raw, flags=re.DOTALL)
        return json.loads(raw)
    except Exception as e:
        logger.error(f"GPT recipe generation failed: {e}")
        return None


async def _gpt_import_recipe_url(url: str) -> dict | None:
    """Fetch a URL and extract a recipe via GPT."""
    from openai import AsyncOpenAI
    import urllib.request
    client = AsyncOpenAI(api_key=OPENAI_API_KEY)
    try:
        req  = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
        # Strip tags roughly
        text = re.sub(r'<[^>]+>', ' ', html)
        text = re.sub(r'\s+', ' ', text)[:6000]
    except Exception as e:
        logger.error(f"URL fetch failed: {e}")
        return None

    prompt = f"""Extract the recipe from this web page text and return ONLY valid JSON with:
name, category, prep_min (int), cook_min (int), servings (int),
calories (int per serving), protein_g (int), carbs_g (int), fat_g (int),
ingredients (multiline string), instructions (numbered multiline string), tags (comma list)

Page text:
{text}"""
    try:
        resp = await client.chat.completions.create(
            model=GPT_CHAT_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
        )
        raw  = resp.choices[0].message.content.strip()
        raw  = re.sub(r'^```json\s*|\s*```$', '', raw, flags=re.DOTALL)
        return json.loads(raw)
    except Exception as e:
        logger.error(f"GPT URL import failed: {e}")
        return None


async def _gpt_suggest_meal_plan(days: int = 7) -> dict | None:
    """Generate a meal plan from the recipe library via GPT."""
    from openai import AsyncOpenAI
    client  = AsyncOpenAI(api_key=OPENAI_API_KEY)
    recipes = _load_recipes()
    if not recipes:
        return None

    recipe_names = ", ".join(str(r.get("name", "")) for r in recipes[:40])
    today = datetime.date.today()
    days_list = [(today + datetime.timedelta(days=i)).isoformat() for i in range(days)]

    prompt = f"""Create a {days}-day meal plan using these available recipes: {recipe_names}

Return ONLY valid JSON — a dict keyed by date ISO string. Each value has keys:
breakfast, lunch, dinner, snacks (all strings, can be empty)

Dates to plan: {', '.join(days_list)}

Vary meals, don't repeat the same dinner twice in a row. Return raw JSON only."""
    try:
        resp = await client.chat.completions.create(
            model=GPT_CHAT_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.8,
        )
        raw  = resp.choices[0].message.content.strip()
        raw  = re.sub(r'^```json\s*|\s*```$', '', raw, flags=re.DOTALL)
        return json.loads(raw)
    except Exception as e:
        logger.error(f"GPT meal plan failed: {e}")
        return None


async def _gpt_nutrition_summary(date_iso: str) -> str:
    """Compute nutrition totals from today's plan."""
    plan = _get_plan_for_date(date_iso)
    if not plan:
        return "No meal plan found for today."

    recipes = _load_recipes()
    recipe_map = {str(r.get("name", "")).lower(): r for r in recipes}
    totals = {"calories": 0, "protein_g": 0, "carbs_g": 0, "fat_g": 0}
    found_any = False

    for slot in ("breakfast", "lunch", "dinner", "snacks"):
        meal_name = str(plan.get(slot, "")).lower()
        if meal_name and meal_name in recipe_map:
            r = recipe_map[meal_name]
            for k in totals:
                try:
                    totals[k] += int(r.get(k) or 0)
                except (TypeError, ValueError):
                    pass
            found_any = True

    if not found_any:
        return "I don't have nutrition data for today's meals. Add macros to your recipes first."

    return (
        f"📊 *Nutrition for {date_iso}*\n"
        f"  🔥 Calories: {totals['calories']} kcal\n"
        f"  💪 Protein:  {totals['protein_g']}g\n"
        f"  🌾 Carbs:    {totals['carbs_g']}g\n"
        f"  🥑 Fat:      {totals['fat_g']}g"
    )


# ─────────────────────────────────────────────────────────────────────────────
# BRIEFING HELPER (called by briefing.py)
# ─────────────────────────────────────────────────────────────────────────────

async def get_todays_meals_text() -> str:
    """Return a one-line meal plan summary for today's briefing."""
    today = _today_iso()
    plan  = _get_plan_for_date(today)
    if not plan:
        return ""
    parts = []
    for slot in ("breakfast", "lunch", "dinner"):
        val = plan.get(slot)
        if val:
            parts.append(f"{slot.title()}: {val}")
    if not parts:
        return ""
    return "🍽 *Today's Meals*\n  " + "  |  ".join(parts)


# ─────────────────────────────────────────────────────────────────────────────
# /meals COMMAND
# ─────────────────────────────────────────────────────────────────────────────

async def cmd_meals(update, context) -> None:
    """Show today's meal plan."""
    today = _today_iso()
    plan  = _get_plan_for_date(today)
    if not plan:
        await update.message.reply_text(
            f"No meal plan for today. Say:\n"
            f"• \"Plan my meals for the week\"\n"
            f"• \"Breakfast today: oats\" to set individual meals\n"
            f"• /meals recipe [name] to look up a recipe",
        )
        return
    await update.message.reply_text(_format_day_plan(plan), parse_mode="Markdown")


# ─────────────────────────────────────────────────────────────────────────────
# INTENT HANDLER
# ─────────────────────────────────────────────────────────────────────────────

async def handle_meal_intent(intent: str, entities: dict, update, context) -> None:
    msg = update.message

    # ── MEAL_VIEW ────────────────────────────────────────────────────────────
    if intent == MEAL_VIEW:
        date_str = entities.get("date", "today")
        if date_str in ("today", ""):
            date_iso = _today_iso()
        else:
            try:
                date_iso = datetime.date.fromisoformat(date_str).isoformat()
            except ValueError:
                date_iso = _today_iso()
        plan = _get_plan_for_date(date_iso)
        if plan:
            await msg.reply_text(_format_day_plan(plan), parse_mode="Markdown")
        else:
            await msg.reply_text(f"No meal plan found for {date_iso}.")
        return

    # ── MEAL_PLAN (set a plan) ────────────────────────────────────────────────
    if intent == MEAL_PLAN:
        action = entities.get("action", "view")
        if action == "view":
            date_iso = _today_iso()
            plan = _get_plan_for_date(date_iso)
            if plan:
                await msg.reply_text(_format_day_plan(plan), parse_mode="Markdown")
            else:
                await msg.reply_text("No meal plan for today. Say \"plan my meals for the week\" to generate one.")
            return

        meals_dict = entities.get("meals", {})
        date_str   = entities.get("date", "today")
        if date_str in ("today", ""):
            date_iso = _today_iso()
        else:
            try:
                date_iso = datetime.date.fromisoformat(date_str).isoformat()
            except ValueError:
                date_iso = _today_iso()

        if not meals_dict:
            # GPT suggestion from library
            await msg.reply_text("⏳ Generating meal plan from your recipe library...")
            plan_dict = await _gpt_suggest_meal_plan(days=7)
            if not plan_dict:
                await msg.reply_text("No recipes in your library yet. Add some recipes first!")
                return
            saved = 0
            for d, m in plan_dict.items():
                if _set_plan_for_date(d, m):
                    saved += 1
            await msg.reply_text(
                f"✓ Meal plan set for {saved} days. Send /meals to see today's plan.",
                parse_mode="Markdown",
            )
        else:
            if _set_plan_for_date(date_iso, meals_dict):
                await msg.reply_text(f"✓ Meal plan saved for {date_iso}.")
            else:
                await msg.reply_text("Couldn't save that plan. Try again.")
        return

    # ── MEAL_RECIPE ───────────────────────────────────────────────────────────
    if intent == MEAL_RECIPE:
        name   = entities.get("name", "").strip()
        if not name:
            recipes = _load_recipes()
            if not recipes:
                await msg.reply_text("No recipes in your library yet.")
                return
            names = "\n".join(f"• {r.get('name','')}" for r in recipes[:20])
            await msg.reply_text(f"*Your Recipes:*\n{names}", parse_mode="Markdown")
            return
        recipe = _find_recipe(name)
        if recipe:
            await msg.reply_text(_format_recipe(recipe), parse_mode="Markdown")
        else:
            await msg.reply_text(
                f"No recipe found for \"{name}\". Say \"generate a recipe for {name}\" to create one."
            )
        return

    # ── MEAL_ADD (save a recipe manually) ─────────────────────────────────────
    if intent == MEAL_ADD:
        name = entities.get("name", "").strip()
        if not name:
            await msg.reply_text("What should I call this recipe?")
            return
        recipe = {k: entities.get(k, "") for k in _RECIPE_COLS}
        recipe["name"] = name
        if _save_recipe(recipe):
            await msg.reply_text(f"✓ Recipe \"{name}\" saved to your library.")
        else:
            await msg.reply_text("Couldn't save that recipe. Try again.")
        return

    # ── MEAL_GENERATE (GPT recipe) ────────────────────────────────────────────
    if intent == MEAL_GENERATE:
        description = entities.get("description", "").strip()
        if not description:
            await msg.reply_text("What should I create a recipe for? e.g. \"generate a high-protein pasta recipe\"")
            return
        await msg.reply_text(f"⏳ Creating a recipe for \"{description}\"...")
        recipe = await _gpt_generate_recipe(description)
        if not recipe:
            await msg.reply_text("Couldn't generate that recipe. Try again.")
            return
        preview = _format_recipe(recipe)
        save    = entities.get("save", True)
        if save:
            _save_recipe(recipe)
            await msg.reply_text(preview + "\n\n✓ Saved to your recipe library.", parse_mode="Markdown")
        else:
            await msg.reply_text(preview + "\n\nSay \"save this recipe\" to add it to your library.", parse_mode="Markdown")
        return

    # ── MEAL_IMPORT (URL) ──────────────────────────────────────────────────────
    if intent == MEAL_IMPORT:
        url = entities.get("url", "").strip()
        if not url:
            await msg.reply_text("Which URL should I import from?")
            return
        await msg.reply_text("⏳ Importing recipe from that URL...")
        recipe = await _gpt_import_recipe_url(url)
        if not recipe:
            await msg.reply_text("Couldn't import from that URL. Try pasting the recipe text instead.")
            return
        _save_recipe(recipe)
        await msg.reply_text(
            f"✓ Imported *{recipe.get('name', 'Recipe')}* and saved to your library.",
            parse_mode="Markdown",
        )
        return

    # ── MEAL_NUTRITION ────────────────────────────────────────────────────────
    if intent == MEAL_NUTRITION:
        date_str = entities.get("date", "today")
        date_iso = _today_iso() if date_str in ("today", "") else date_str
        summary  = await _gpt_nutrition_summary(date_iso)
        await msg.reply_text(summary, parse_mode="Markdown")
        return

    # ── MEAL_ADHERENCE ────────────────────────────────────────────────────────
    if intent == MEAL_ADHERENCE:
        date_iso = _today_iso()
        plan     = _get_plan_for_date(date_iso)
        notes    = entities.get("notes", "")
        if not plan:
            await msg.reply_text("No meal plan found for today to log adherence against.")
            return
        _log_adherence(date_iso, "all", str(plan), notes)
        await msg.reply_text(f"✓ Meal adherence logged for {date_iso}.")
        return

    # ── MEAL_EXPORT ───────────────────────────────────────────────────────────
    if intent == MEAL_EXPORT:
        _ensure_workbook()
        if not os.path.exists(MEALS_XLSX):
            await msg.reply_text("No meal data yet.")
            return
        await context.bot.send_document(
            chat_id=msg.chat_id,
            document=open(MEALS_XLSX, "rb"),
            filename="meals.xlsx",
            caption="Your meal plan and recipe library.",
        )
        return

    # ── MEAL_LEFTOVERS ────────────────────────────────────────────────────────
    if intent == MEAL_LEFTOVERS:
        action  = entities.get("action", "log")
        details = entities.get("details", "")
        from core.data import load_data, save_data
        data = load_data()
        data.setdefault("settings", {}).setdefault("leftovers", {})
        today = _today_iso()
        if action == "log":
            data["settings"]["leftovers"][today] = details
            save_data(data)
            await msg.reply_text(f"✓ Leftovers logged: {details}")
        else:
            leftovers = data["settings"].get("leftovers", {})
            if leftovers:
                recent = sorted(leftovers.items())[-5:]
                lines  = [f"  {d}: {v}" for d, v in reversed(recent)]
                await msg.reply_text("🍱 *Recent Leftovers:*\n" + "\n".join(lines), parse_mode="Markdown")
            else:
                await msg.reply_text("No leftovers logged yet.")
        return


# ─────────────────────────────────────────────────────────────────────────────
# SCHEDULED: meal adherence check
# ─────────────────────────────────────────────────────────────────────────────

async def send_meal_adherence_check(context, chat_id: int) -> None:
    """Evening prompt: did you eat as planned?"""
    today = _today_iso()
    plan  = _get_plan_for_date(today)
    if not plan:
        return  # No plan today, skip

    planned_meals = []
    for slot in ("breakfast", "lunch", "dinner"):
        val = plan.get(slot)
        if val:
            planned_meals.append(f"  {slot.title()}: {val}")

    if not planned_meals:
        return

    lines = ["🍽 *Meal check-in*", "Today's plan:"] + planned_meals + [
        "",
        "Did you eat as planned? Reply with:",
        "• \"Yes, I ate as planned\"",
        "• \"I had [something different] for [meal slot]\"",
        "• \"Skip\" to log it later",
    ]
    await context.bot.send_message(
        chat_id=chat_id,
        text="\n".join(lines),
        parse_mode="Markdown",
    )
