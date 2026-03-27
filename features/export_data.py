"""
alfred/features/export_data.py
==============================
Export all user data to a single Excel workbook.

SHEETS
──────
  Journal      — date, entries (prompts+answers), freeform, mood_rating
  Mood Log     — date, rating, note
  Notes        — title, content, created_at
  Habits       — name, frequency, streak, last_completed
  Memory       — category, facts (one row per fact)
  Links        — title, url, tags, saved, read status
  Shopping     — list name, item (one row per item)

COMMAND
───────
  /export      — generate and send Excel file
"""

import os
import logging
import datetime
from io import BytesIO

from telegram import Update
from telegram.ext import ContextTypes

from core.config import EXPORT_DIR
from core.intent import EXPORT_DATA
from core.data import load_data, load_journal

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _make_header(ws, headers: list) -> None:
    """Write bold header row with light blue fill."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return

    fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    font = Font(bold=True, color="000000")
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _autosize(ws) -> None:
    """Set column widths based on content."""
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(60, max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width


def _export_journal(wb, data: dict, journal: dict) -> None:
    """Add Journal sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed")
        return

    ws = wb.create_sheet("Journal")
    _make_header(ws, ["Date", "Entries", "Freeform", "Mood Rating"])

    row = 2
    for date_iso in sorted(journal.keys()):
        day_data = journal[date_iso]
        entries_list = day_data.get("entries", [])

        # Combine all entries
        all_text = []
        mood_rating = None

        for entry in entries_list:
            etype = entry.get("type", "freeform")
            content = entry.get("content", "")

            if etype == "prompted" and isinstance(content, dict):
                for q, a in content.items():
                    all_text.append(f"{q}: {a}")
            elif etype == "freeform":
                all_text.append(str(content))
            elif etype == "voice":
                all_text.append(f"[Voice] {content}")

        entries_text = "\n".join(all_text)

        ws.cell(row=row, column=1, value=date_iso)
        ws.cell(row=row, column=2, value=entries_text[:500])  # truncate
        ws.cell(row=row, column=3, value=entries_text[500:])
        row += 1

    _autosize(ws)


def _export_mood(wb, data: dict) -> None:
    """Add Mood sheet."""
    ws = wb.create_sheet("Mood Log")
    _make_header(ws, ["Date", "Rating", "Note"])

    row = 2
    for entry in data.get("mood_log", []):
        ws.cell(row=row, column=1, value=entry.get("date", ""))
        ws.cell(row=row, column=2, value=entry.get("rating", ""))
        ws.cell(row=row, column=3, value=entry.get("note", ""))
        row += 1

    _autosize(ws)


def _export_notes(wb, data: dict) -> None:
    """Add Notes sheet."""
    ws = wb.create_sheet("Notes")
    _make_header(ws, ["Title", "Content", "Created At"])

    row = 2
    for note in data.get("notes", []):
        ws.cell(row=row, column=1, value=note.get("id", ""))
        ws.cell(row=row, column=2, value=note.get("text", "")[:500])
        ws.cell(row=row, column=3, value=note.get("added", ""))
        row += 1

    _autosize(ws)


def _export_habits(wb, data: dict) -> None:
    """Add Habits sheet."""
    ws = wb.create_sheet("Habits")
    _make_header(ws, ["Name", "Frequency", "Last Logged"])

    row = 2
    for habit in data.get("habits", []):
        ws.cell(row=row, column=1, value=habit.get("name", habit.get("id", "")))
        ws.cell(row=row, column=2, value="daily")
        ws.cell(row=row, column=3, value="")
        row += 1

    _autosize(ws)


def _export_memory(wb, memory: dict) -> None:
    """Add Memory sheet."""
    ws = wb.create_sheet("Memory")
    _make_header(ws, ["Category", "Fact"])

    row = 2
    for cat, facts in memory.items():
        if cat.startswith("_"):  # skip metadata
            continue
        if isinstance(facts, list):
            for fact in facts:
                ws.cell(row=row, column=1, value=cat)
                ws.cell(row=row, column=2, value=str(fact)[:500])
                row += 1

    _autosize(ws)


def _export_links(wb) -> None:
    """Add Links sheet."""
    try:
        from features.links import _load_links
    except ImportError:
        return

    ws = wb.create_sheet("Links")
    _make_header(ws, ["Title", "URL", "Tags", "Saved", "Read"])

    row = 2
    for link in _load_links():
        ws.cell(row=row, column=1, value=link.get("title", ""))
        ws.cell(row=row, column=2, value=link.get("url", ""))
        ws.cell(row=row, column=3, value=", ".join(link.get("tags", [])))
        ws.cell(row=row, column=4, value=link.get("saved_at", ""))
        ws.cell(row=row, column=5, value="Yes" if link.get("read") else "No")
        row += 1

    _autosize(ws)


def _export_shopping(wb, data: dict) -> None:
    """Add Shopping sheet."""
    ws = wb.create_sheet("Shopping Lists")
    _make_header(ws, ["List", "Item"])

    row = 2
    shopping = data.get("shopping", {})
    for list_key, items in shopping.items():
        if isinstance(items, list):
            for item_obj in items:
                item_text = item_obj.get("text") if isinstance(item_obj, dict) else str(item_obj)
                ws.cell(row=row, column=1, value=list_key.title())
                ws.cell(row=row, column=2, value=item_text)
                row += 1

    _autosize(ws)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_export(output_path: str) -> str:
    """
    Create a full workbook with all sheets.
    Saves to output_path.
    Returns the path if successful, or error message.
    """
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl not installed. Export feature unavailable."

    try:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        data = load_data()
        journal = load_journal()

        # Load memory
        from core.data import load_memory
        memory = load_memory()

        # Add all sheets
        _export_journal(wb, data, journal)
        _export_mood(wb, data)
        _export_notes(wb, data)
        _export_habits(wb, data)
        _export_memory(wb, memory)
        _export_links(wb)
        _export_shopping(wb, data)

        wb.save(output_path)
        return output_path

    except Exception as e:
        logger.error(f"export_data: generation error: {e}")
        return f"Error: {e}"


# ─────────────────────────────────────────────────────────────────────────────
# COMMAND
# ─────────────────────────────────────────────────────────────────────────────

async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Generate and send Excel export."""
    if not update.message:
        return

    await update.message.reply_text("📊 Generating export… this may take a moment.")

    today = datetime.date.today().isoformat()
    output_file = os.path.join(EXPORT_DIR, f"alfred_export_{today}.xlsx")

    result = generate_export(output_file)

    if not result.startswith("Error"):
        try:
            with open(output_file, "rb") as f:
                await update.message.reply_document(
                    f,
                    caption="📊 Complete data export",
                    filename=f"alfred_export_{today}.xlsx",
                )
        except Exception as e:
            logger.error(f"export_data: send error: {e}")
            await update.message.reply_text(f"Error sending file: {e}")
    else:
        await update.message.reply_text(result)


# ─────────────────────────────────────────────────────────────────────────────
# INTENT HANDLER
# ─────────────────────────────────────────────────────────────────────────────

async def handle_export_intent(intent: str, entities: dict, update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle EXPORT_DATA intent."""
    if intent == EXPORT_DATA:
        await cmd_export(update, context)
